import openpyxl

# Eurostat file, downloaded from: https://ec.europa.eu/eurostat/databrowser/view/road_go_ca_hac/default/table?lang=en
file_path = "../data/raw/road_go_ca_hac_spreadsheet.xlsx"
workbook = openpyxl.load_workbook(file_path)
country_data = {}


def process_sheet(sheet):
    country_of_interest = sheet["C7"].value.lower()

    data = []
    for row in sheet.iter_rows(min_row=12, max_row=41, min_col=1, max_col=49):
        country = row[0].value
        tonnage = row[47].value

        if tonnage != ":":
            data.append((country, float(tonnage)))

    if data:
        # sort the data by tonnage
        sorted_data = sorted(data, key=lambda x: x[1], reverse=True)

        # get the top 10 performing countries
        top = sorted_data[:10]

        # calculate total tonnage for all countries
        total_tonnage = sum(entry[1] for entry in sorted_data)

        # make a dictionary with the top 10, and the corresponding percentages
        country_data[country_of_interest] = {}
        for entry in top:
            percentage = round((entry[1] / total_tonnage), ndigits=2)
            country_data[country_of_interest][entry[0].lower()] = percentage


# process each sheet
for sheet_name in workbook.sheetnames:
    if sheet_name.startswith("Sheet") and float(sheet_name.split(" ")[1]) <= 40:
        sheet = workbook[sheet_name]
        process_sheet(sheet)

for key in country_data:
    print(f"\n{key}\n{country_data[key]}\n")


def retrieve_scaled(scale_factor, country):
    sum = 0
    scaled_dict = {}
    for foreign_country in country_data[country]:
        sum += country_data[country][foreign_country]
    for foreign_country in country_data[country]:
        scaled_dict[foreign_country] = round(country_data[country][foreign_country] / sum * scale_factor, 2)
    return scaled_dict


def print_scaled_dict(scaled_dict):
    res_string = "["
    for index, country in enumerate(scaled_dict.keys()):
        res_string += f"(\"{country}\", {scaled_dict[country]})"
        if index < len(scaled_dict.keys()) - 1:
            res_string += ", "
    res_string += "]"
    print(res_string)


print_scaled_dict(retrieve_scaled(0.3, "slovakia"))
